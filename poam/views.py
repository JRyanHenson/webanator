from django.shortcuts import render, redirect, reverse
from django.views import generic
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
from xml.etree import ElementTree
from django.db import IntegrityError

import openpyxl
import datetime
# import os

from .forms import *
from .models import *


# Create your views here.
class UploadArtifactView(LoginRequiredMixin, generic.CreateView):
    template_name = 'poam/upload_artifact.html'
    model = Weakness
    form_class = DocumentForm

    def get(self, request, *args, **kwargs):
        system = System.objects.get(id=self.kwargs['pk'])
        form = self.form_class(initial=self.initial, system=system)
        return render(request, self.template_name, {'form': form, 'system': system})

    def get_form_kwargs(self):
        kwargs = super(UploadArtifactView, self).get_form_kwargs()
        kwargs['system'] = System.objects.get(id=self.kwargs['pk'])
        return kwargs

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        files = request.FILES.getlist('file')
        if form.is_valid():
            form.cleaned_data['file'] = []
            for f in files:
                form.cleaned_data['file'].append(f)
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        source_id = form.cleaned_data['source_identifying_event']
        source_tool = form.cleaned_data['source_identifying_tool']
        source_date = form.cleaned_data['source_identifying_date']
        system = System.objects.get(id=self.kwargs['pk'])
        poc = form.cleaned_data['point_of_contact']
        file_type = form.cleaned_data['file_type']
        device_queryset = form.cleaned_data['devices']
        for data in form.cleaned_data['file']:
            # read and parse the file, create a Python dictionary `data_dict` from it
            # start loop here for each Vuln_num in xml upload get Rule_title, Vuln_discuss,
            # Comments, IA_Controls, Check_Content, Severity, Stigid
            # Need check and message to make sure Trey or more than likely Zac doesn't do something dumb
            if file_type == 'stig_checklist_file':
                tree = ElementTree.parse(data)
                root = tree.getroot()
                if root.tag == "CHECKLIST":
                    for vuln in root.findall('.//VULN'):
                        status = vuln.find('STATUS').text
                        comments = vuln.find('COMMENTS').text
                        finding_details = vuln.find('FINDING_DETAILS').text
                        severity = vuln[1][1].text
                        title = vuln[5][1].text
                        description = vuln[6][1].text
                        chk_content = vuln[8][1].text
                        fix_text = vuln[9][1].text
                        stig_ref = vuln[22][1].text
                        vuln_id = vuln[0][1].text
                        diacap_ia_control = None
                        try:
                            vuln_id = VulnId(vuln_id=vuln_id)
                            vuln_id.save()
                        except IntegrityError:
                            vuln_id = VulnId.objects.get(vuln_id=vuln_id.vuln_id)
                        try:
                            cci = vuln[24][1].text
                            cci = CCI.objects.get(cci=cci)
                        except:
                            diacap_ia_control = vuln[7][1].text
                            cci = None
                        if status == 'Open':
                            for device in device_queryset:
                                if not Device.objects.filter(system=system, name=device).exists():
                                    messages.error(self.request, 'Woops! One of the devices is not in the database')
                                else:
                                    device = Device.objects.get(system=system, name=device)
                                    if Weakness.objects.filter(devices__id=device.id, vuln_id=vuln_id.id).exists():
                                        Weakness.objects.filter(system=system, vuln_id=vuln_id.id).update(comments=comments, finding_details=finding_details,
                                        stig_ref=stig_ref, raw_severity=severity, source_identifying_event=source_id,
                                        source_identifying_tool=source_tool, check_contents=chk_content, fix_text=fix_text,
                                        point_of_contact=PointOfContact.objects.get(name=poc).id, status=status, security_control=diacap_ia_control)
                                        if cci is not None:
                                            Weakness.objects.filter(system=system, vuln_id=vuln_id.id).update(cci=cci.id)
                                    else:
                                        data_dict = {'title': title, 'description': description, 'status': status,
                                                     'comments': comments, 'finding_details': finding_details, 'stig_ref': stig_ref,'raw_severity': severity,
                                                     'source_identifying_event': source_id, 'source_identifying_tool': source_tool, 'source_identifying_date': source_date,
                                                     'vuln_id': vuln_id.id, 'check_contents': chk_content, 'fix_text': fix_text, 'system': system.id,
                                                     'point_of_contact': poc.id, 'devices': device_queryset, 'security_control': diacap_ia_control}
                                        if cci is not None:
                                            data_dict['cci'] = cci.id
                                        form = WeaknessModelForm(data_dict)
                                        try:
                                            form.save()
                                        except:
                                            messages.error(self.request, '{} had error when saving to database. {}'.format(vuln_id, form.errors))
                        else:
                            continue
                else:
                    messages.error(self.request, 'Wrong File Type, Zac!')
                    return redirect(reverse("poam:upload-artifact"))
            elif file_type == 'rmf_control_review_file':
                try:
                    wb = openpyxl.load_workbook(data)
                    sheet = wb.active
                except:
                    messages.error(self.request, 'Wrong File Type, Zac!')
                    return redirect(reverse("poam:upload-artifact"))
                else:
                    for row in range(2, sheet.max_row + 1):
                        title = sheet['B{}'.format(row)].value
                        status = sheet['C{}'.format(row)].value
                        comments = sheet['D{}'.format(row)].value
                        control_number = sheet['A{}'.format(row)].value
                        for device in device_queryset:
                            if not Device.objects.filter(system=system, name=device).exists():
                                messages.error(self.request, 'Woops! One of the devices is not in the database')
                            else:
                                device = Device.objects.get(system=system, name=device)
                            if control_number is not None:
                                try:
                                    security_control = SecurityControl(title=title, control_number=control_number, description='')
                                    security_control.save()
                                except IntegrityError:
                                    security_control = SecurityControl.objects.get(control_number=control_number)
                                    print(security_control.control_number)
                                try:
                                    vuln_id = VulnId(vuln_id=security_control.control_number)
                                    vuln_id.save()
                                except IntegrityError:
                                    vuln_id = VulnId.objects.get(vuln_id=security_control.control_number)

                                if Weakness.objects.filter(title=security_control.title, devices=device.id).exists():
                                    print('hi')
                                    print(vuln_id.vuln_id)
                                    Weakness.objects.filter(title=security_control.title, devices=device.id).update(status=status, comments=comments, source_identifying_event=source_id, source_identifying_tool=source_tool, security_control=security_control.control_number, source_identifying_date=source_date, system=System.objects.get(name=system).id, point_of_contact=PointOfContact.objects.get(name=poc).id, vuln_id=vuln_id)
                                else:
                                    continue
                                    # if status == 'Planned':
                                    #     data_dict = {'title': title, 'description': '', 'status': status, 'comments': comments, 'source_identifying_event': source_id, 'source_identifying_tool': source_tool, 'devices': device_queryset, 'security_control': security_control, 'source_identifying_date' : source_date, 'system': System.objects.get(name=system).id, 'point_of_contact': PointOfContact.objects.get(name=poc).id, 'vuln_id' : security_control}
                                    #     form = WeaknessModelForm(data_dict)
                                    #     try:
                                    #         form.save()
                                    #     except:
                                    #         messages.error(self.request, form.errors)
            elif file_type == 'nessus_scan_file':
                # parse data from nessus file and define as tree then get the root of the xml file
                tree = ElementTree.parse(data)
                root = tree.getroot()
                # verify that the file is in fact a nessus file by checking the root tag of the xml
                if root.tag == 'NessusClientData_v2':
                    # define static variables from nessus file
                    # hostname = root.findtext(".//tag[@name='hostname']")
                    ip = root.findtext(".//tag[@name='host-ip']")
                    os = root.findtext(".//tag[@name='operating-system']")
                    hostname = root.findtext(".//tag[@name='hostname']")
                    netbios_name = root.findtext(".//tag[@name='netbios-name']")
                    mac = root.findtext(".//tag[@name='mac-address']")
                    bios_uid = root.findtext(".//tag[@name='bios-uuid']")
                    system_cpe = [root.findtext(".//tag[@name='cpe']"), root.findtext(".//tag[@name='cpe-0']"), root.findtext(".//tag[@name='cpe-1']"), root.findtext(".//tag[@name='cpe-2']"), root.findtext(".//tag[@name='cpe-3']"), root.findtext(".//tag[@name='cpe-4']"), root.findtext(".//tag[@name='cpe-5']")]
                    credentialed_scan = root.findtext(".//tag[@name='Credentialed_Scan']")
                    for device in device_queryset:
                        if not Device.objects.filter(system=system, name=device).exists():
                            messages.error(self.request, 'Woops! One of the devices is not in the database')
                            break
                        else:
                            Device.objects.filter(system=system, name=device).update(os=os, ip=ip, hostname=hostname, netbios_name=netbios_name, mac=mac, bios_uid=bios_uid)
                    # for loop of nessus xml file to get relevant weakness information
                    for cpe in system_cpe:
                        if cpe is not None:
                            try:
                                cpe = CPE(cpe=cpe)
                                cpe.save()
                            except IntegrityError:
                                cpe = CPE.objects.get(cpe=cpe.cpe)
                            if not Device.objects.filter(system=system, name=device, cpes=cpe.id).exists():
                                device = Device.objects.get(system=system, name=device)
                                device.cpes.add(cpe)
                                device.save()
                    for vuln in tree.iter('ReportItem'):
                        vuln_id = vuln.get('pluginID')
                        severity = vuln.get('severity')
                        description = vuln.findtext('description')
                        title = vuln.get('pluginName')
                        plugin_family = vuln.get('pluginFamily')
                        plugin_output = vuln.findtext('plugin_output')
                        fix_text = vuln.findtext('solution')
                        synopsis = vuln.findtext('synopsis')
                        status = 'open'
                        cvss_base_score = vuln.findtext('cvss_base_score')
                        cvss_temporal_score = vuln.findtext('cvss_temporal_score')
                        cvss_vector = vuln.findtext('cvss_vector')
                        cvss_temporal_vector = vuln.findtext('cvss_temporal_vector')
                        cvss3_base_score = vuln.findtext('cvss3_base_score')
                        cvss3_vector = vuln.findtext('cvss3_vector')
                        exploit_available = vuln.findtext('exploit_available')
                        findings_cpe = vuln.findtext('cpe')
                        cve = vuln.findtext('cve')
                        risk_factor = vuln.findtext('risk_factor')
                        vuln_pub_date = vuln.findtext('vuln_publication_date')
                        # try block to test to see if vuln_id already exists in DB. if not, creates object
                        if findings_cpe is not None:
                            try:
                                findings_cpe = CPE(cpe=findings_cpe)
                                findings_cpe.save()
                            except IntegrityError:
                                findings_cpe = CPE.objects.get(cpe=findings_cpe.cpe)
                            if not Device.objects.filter(name=device, cpes=findings_cpe.id).exists():
                                device = Device.objects.get(system=system, name=device)
                                device.cpes.add(findings_cpe)
                                device.save()
                        if severity != '0':
                            try:
                                vuln_id = VulnId(vuln_id=vuln_id)
                                vuln_id.save()
                            except IntegrityError:
                                vuln_id = vuln.get('pluginID')
                                vuln_id = VulnId.objects.get(vuln_id=vuln_id)
                            # checks to see if weakness object with this hostname and vuln id already exists. if so, it updates existing object
                            if Weakness.objects.filter(devices__id=device.id, vuln_id=vuln_id.id).exists():
                                Weakness.objects.filter(devices__id=device.id, vuln_id=vuln_id.id).update(raw_severity=severity,
                                 plugin_family=plugin_family, synopsis=synopsis, plugin_output=plugin_output,source_identifying_event=source_id, source_identifying_tool=source_tool,
                                 fix_text=fix_text, point_of_contact=PointOfContact.objects.get(name=poc), status=status, cvss_base_score=cvss_base_score,
                                 cvss_temporal_score=cvss_temporal_score, cvss_vector=cvss_vector, cvss_temporal_vector=cvss_temporal_vector, cvss3_base_score=cvss3_base_score, cvss3_vector=cvss3_vector,
                                 cve=cve, risk_factor=risk_factor, vuln_pub_date=vuln_pub_date, exploit_available=exploit_available,
                                 credentialed_scan=credentialed_scan)
                            else:
                                # if severity is not 0 then create and new weakness object using a data dictionary file
                                weakness_data_dict = {'title': title, 'system': system.id, 'description': description, 'status': status,
                                             'raw_severity': severity, 'source_identifying_event': source_id, 'source_identifying_date': source_date,
                                             'source_identifying_tool': source_tool, 'vuln_id': vuln_id.id, 'fix_text': fix_text, 'devices': device_queryset,
                                             'point_of_contact' : poc.id, 'plugin_family': plugin_family, 'plugin_output': plugin_output,
                                             'synopsis': synopsis, 'credentialed_scan': credentialed_scan, 'cvss_base_score': cvss_base_score, 'cvss_vector': cvss_vector,
                                             'cvss_temporal_score': cvss_temporal_score, 'cvss_temporal_vector': cvss_temporal_vector,
                                             'cvss3_base_score': cvss3_base_score, 'cvss3_vector': cvss3_vector, 'exploit_available': exploit_available, 'cve': cve, 'risk_factor': risk_factor,
                                             'vuln_pub_date': vuln_pub_date}
                                weaknessform = WeaknessModelForm(weakness_data_dict)
                                try:
                                    weaknessform.save()
                                except:
                                    messages.error(self.request, '{} had error when saving to database. {}'.format(vuln_id.vuln_id, form.errors))
                        else:
                            continue
                                    # checks for weakness objects created before the date of the current upload. if date is different,
                                    # updates weakness object to mark previous results as closed. eventually add logic to get user confirmation
                                    # if Weakness.objects.exclude(devices__id=specific_device.id, source_identifying_date=source_date).exists():
                                    #     messages.success(self.request, 'Credentialed Scan: ' + credentialed_scan + ' Previous scan results will be closed')
                                    #     Weakness.objects.exclude(devices__id=specific_device.id, source_identifying_date=source_date).update(status='closed')
                else:
                    messages.error(self.request, 'Wrong File Type, Zac!')
                    weaknessform = WeaknessModelForm()
                    weaknessform.save()
            # elif file_type == 'cci_list':
            # # Used to upload CCI/SecurityControl List. Not needed for regular use.
            #     tree = ElementTree.parse(data)
            #     root = tree.getroot()
            #     if root.tag == "{http://iase.disa.mil/cci}cci_list":
            #         for child in tree.iter('{http://iase.disa.mil/cci}cci_item'):
            #             cci = child.get('id')
            #             definition = child.find('{http://iase.disa.mil/cci}definition').text
            #             references = child.findall(".//*[@title='NIST SP 800-53 Revision 4']")
            #             for reference in references:
            #                 scref = reference.get('index')
            #                 sc = scref.split(" ")[0]
            #                 if not CCI.objects.filter(cci=cci).exists():
            #                     sci_input = CCI(cci=cci, sc=sc, definition=definition, scref=scref)
            #                     sci_input.save()
            #                 else:
            #                     continue
            #     else:
            #         messages.error(self.request, 'Wrong File Type, Zac!')
            #         return redirect(reverse("poam:upload-artifact"))
            elif file_type == 'poam_format':
                try:
                    wb = openpyxl.load_workbook(data)
                    sheet = wb.active
                except:
                    messages.error(self.request, 'Wrong File Type, Zac!')
                    return redirect(reverse("poam:upload-artifact"))
                else:
                    # system values
                    system_name = sheet['B4'].value
                    system_update_date = sheet['B2'].value
                    system_create_date = sheet['B1'].value
                    system_dod_component = sheet['B3'].value
                    system_dod_it_resource_number = sheet['B5'].value
                    system_type = sheet['H1'].value
                    system_poc_name = sheet['H3'].value
                    system_poc_email = sheet['H4'].value
                    system_poc_phone = sheet['H5'].value

                    # check if point of contact exists in database
                    # if not, create new one
                    # if yes, set poc to point of contact
                    try:
                        poc = PointOfContact.objects.get(name=system_poc_name)
                    except PointOfContact.DoesNotExist:
                        poc = PointOfContact.objects.create(name=system_poc_name, email=system_poc_email, phone=system_poc_phone)

                    # check if system exists in database
                    # if not, possible form tampering, break
                    # else, update system information
                    try:
                        system = System.objects.get(name=system_name)
                    except System.DoesNotExist:
                        messages.error(self.request, 'System names don\'t match! Are you sure you chose the correct system?')
                        break
                    else:
                        System.objects.filter(name=system_name).update(update_date=system_update_date, dod_component=system_dod_component, dod_it_resource_number=system_dod_it_resource_number, system_type=system_type, point_of_contact=poc)

                        system = System.objects.get(name=system_name)
                        for row in range(7, sheet.max_row + 1):
                            weakness = sheet['A{}'.format(row)].value

                            # Gets everything before Title: - vuln_id
                            vuln_id = weakness.split("Title: ")[0]
                            vuln_id = vuln_id.rstrip()

                            # Gets everything after Title:
                            weakness = weakness.split("Title: ")[1]

                            # Gets everything before Description: - title
                            title = weakness.split("Description: ")[0]
                            title = title.rstrip()

                            # Gets everything after Description
                            weakness = weakness.split("Description: ")[1]

                            if 'Devices Affected:' in weakness:
                                # Gets everything after Devices Affected:
                                devices_affected = weakness.split("Devices Affected:\n")[1]
                                devices_affected = devices_affected.rstrip()
                                devices_affected = devices_affected.rsplit('\n')

                                # Gets everything before Devices Affected: - Description
                                weakness = weakness.split("Devices Affected:")[0]
                                weakness = weakness.rstrip()

                                #

            # want to close nessus weakness from previous scans on this device
            # if Weakness.objects.exclude(system=system, source_identifying_date=source_date).exists():
            #     messages.success(self.request, 'Credentialed Scan: {} Would you like to close previous ACAS scan results for this device?'.format(credentialed_scan))
            #     Weakness.objects.exclude(system=system, source_identifying_date=source_date).update(status='closed')
        messages.success(self.request, 'Artifacts Uploaded Successfully!')
        return redirect(reverse('poam:edit-system', kwargs={'pk': self.kwargs['pk']}))


class NewSystemView(LoginRequiredMixin, generic.CreateView):
    template_name= 'poam/new_system.html'
    model = System
    form_class = SystemForm

    def form_valid(self, form):
        self.object=form.save()
        return redirect(reverse('accounts:dashboard'))


class NewPOCView(LoginRequiredMixin, generic.CreateView):
    template_name = 'poam/new_poc.html'
    model = PointOfContact
    form_class = NewPOCForm

    def form_valid(self, form):
        self.object=form.save()
        return redirect(reverse("poam:new-system"))


class SelectSystemView(LoginRequiredMixin, generic.ListView):
    template_name = "poam/select-system.html"
    model = System
    context_object_name = 'systems'

    def post(self, request, *args, **kwargs):
        form = request.POST
        system = form['system']
        return redirect(reverse('poam:edit-system', kwargs={'pk': system}))


class EditSystemView(LoginRequiredMixin, generic.UpdateView):
    template_name = "poam/edit-system.html"
    model = System
    form_class = SystemForm

    def get_success_url(self):
        url = reverse('poam:edit-system', kwargs={'pk': self.kwargs['pk']})
        return url


class AddDeviceView(LoginRequiredMixin, generic.CreateView):
    template_name = "poam/add-device.html"
    model = Device
    form_class = DeviceForm

    def get_context_data(self, **kwargs):
        context = super(AddDeviceView, self).get_context_data(**kwargs)
        context['system'] = System.objects.get(id=self.kwargs['pk'])
        return context

    def form_valid(self, form):
        device = form.save(commit=False)
        device.system = System.objects.get(id=self.kwargs['pk'])
        device.save()
        messages.success(self.request, 'Device Added Successfully!')
        return redirect(reverse('poam:edit-system', kwargs={'pk': self.kwargs['pk']}))


class ExportPoamView(LoginRequiredMixin, generic.DetailView):
    template_name = 'poam/edit-system.html'
    model = System

    def get(self, request, *args, **kwargs):
        fp = openpyxl.load_workbook('{}/template.xlsx'.format(settings.MEDIA_ROOT))
        ws = fp['POAM']

        ws['B1'] = self.get_object().create_date
        ws['B2'] = self.get_object().update_date
        ws['B3'] = self.get_object().dod_component
        ws['B4'] = self.get_object().name
        ws['B5'] = self.get_object().dod_it_resource_number

        ws['H1'] = self.get_object().system_type
        ws['H3'] = self.get_object().point_of_contact.name
        ws['H4'] = self.get_object().point_of_contact.email
        ws['H5'] = self.get_object().point_of_contact.phone

        poams = self.get_object().get_weaknesses()
        row = 7
        vulns = []
        for poam in poams:
            if poam.vuln_id and poam.vuln_id.vuln_id not in vulns:
                weakness = ''
                weakness += '{}\n'.format(poam.vuln_id.vuln_id)
                weakness += 'Title: {}\n'.format(poam.title)
                weakness += '\nDescription: {}'.format(poam.description)
                weakness += '\n\nDevices Affected:'
                for weak in Weakness.objects.filter(vuln_id=poam.vuln_id):
                    for device in Device.objects.filter(weakness=weak.id):
                        weakness += '\n{}'.format(device.name)
                ws['A{}'.format(row)] = weakness

                if poam.raw_severity.lower() == 'high' or poam.raw_severity.lower() == 'very high' or poam.raw_severity == '1':
                    raw_severity = 'I'
                elif poam.raw_severity.lower() == 'medium' or poam.raw_severity == '2':
                    raw_severity = 'II'
                elif poam.raw_severity.lower() == 'low' or poam.raw_severity == '3' or poam.raw_severity == '4':
                    raw_severity = 'III'
                else:
                    raw_severity = poam.raw_severity

                ws['B{}'.format(row)] = raw_severity
                try:
                    ws['C{}'.format(row)] = poam.cci.scref
                except AttributeError:
                    ws['C{}'.format(row)] = poam.security_control
                ws['D{}'.format(row)] = poam.mitigated_severity
                ws['E{}'.format(row)] = poam.mitigation

                poc = ''
                poc += '{}\n'.format(poam.point_of_contact.name)
                poc += '{}\n'.format(poam.point_of_contact.email)
                poc += '{}\n'.format(poam.point_of_contact.phone)

                ws['F{}'.format(row)] = poc
                ws['G{}'.format(row)] = poam.resources_required
                ws['H{}'.format(row)] = poam.scheduled_completion_date
                ws['I{}'.format(row)] = poam.milestone_changes

                source_identifying_weakness = ''
                source_identifying_weakness += '1. {}\n'.format(poam.source_identifying_event)
                stig_ref = poam.stig_ref
                print(stig_ref)
                if stig_ref is not "":
                    source_identifying_weakness += '2. {}\n'.format(poam.stig_ref)
                else:
                    source_identifying_weakness += '2. {}\n'.format(poam.source_identifying_tool)
                source_identifying_weakness += '3. {}\n'.format(poam.source_identifying_date)

                comments = ''
                comments += 'Comments:  {}\n'.format(poam.comments)
                comments += '\nFinding Details:  {}\n'.format(poam.finding_details)
                comments += '\nCVSS Scores:  \n'
                comments += 'CVSS2 Base Score - {}\n'.format(poam.cvss_base_score)
                comments += 'CVSS2 Temporal Score - {}\n'.format(poam.cvss_temporal_score)
                comments += 'CVSS2 Vector - {}\n'.format(poam.cvss_vector)
                comments += 'CVSS2 Temporal Vector - {}\n'.format(poam.cvss_temporal_vector)
                comments += 'CVSS3 Base Score - {}\n'.format(poam.cvss3_base_score)
                comments += 'CVSS3 Vector - {}\n'.format(poam.cvss3_vector)
                ws['J{}'.format(row)] = source_identifying_weakness
                ws['K{}'.format(row)] = poam.status
                ws['L{}'.format(row)] = comments
                row += 1
                vulns.append(poam.vuln_id.vuln_id)
            else:
                continue

        filename = '{}_{}.xlsx'.format(datetime.date.today(), self.get_object().name)
        fp.save('{}/poam/{}'.format(settings.MEDIA_ROOT, filename))

        fp = openpyxl.load_workbook('{}/poam/{}'.format(settings.MEDIA_ROOT, filename))
        httpresponse = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(fp), content_type='application/vnd.ms-excel')
        httpresponse['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        return httpresponse


class ExportHwSwView(LoginRequiredMixin, generic.DetailView):
    template_name = 'poam/edit-system.html'
    model = System

    def get(self, request, *args, **kwargs):
        fp = openpyxl.load_workbook('{}/template2.xlsx'.format(settings.MEDIA_ROOT))
        ws = fp['baseline']

        hwsw = self.get_object().get_devices()
        row = 2
        for hwsw in hwsw:
            ws['A{}'.format(row)] = hwsw.name
            ws['B{}'.format(row)] = hwsw.type
            ws['C{}'.format(row)] = hwsw.hardware
            ws['D{}'.format(row)] = hwsw.hostname
            ws['E{}'.format(row)] = hwsw.ip
            ws['F{}'.format(row)] = hwsw.os
            software = ''
            for cpe in CPE.objects.filter(device=hwsw.id):
                software += '- {}\n'.format(cpe.cpe)
            ws['G{}'.format(row)] = software
            row += 1

        filename = '{}_{}.xlsx'.format(datetime.date.today(), self.get_object().name)
        fp.save('{}/poam/{}'.format(settings.MEDIA_ROOT, filename))

        fp = openpyxl.load_workbook('{}/poam/{}'.format(settings.MEDIA_ROOT, filename))
        httpresponse = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(fp), content_type='application/vnd.ms-excel')
        httpresponse['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        return httpresponse


class SelectDeviceView(LoginRequiredMixin, generic.ListView):
    template_name = 'poam/select-device.html'
    model = Device
    context_object_name = 'devices'

    def get(self, request, *args, **kwargs):
        system = System.objects.get(id=self.kwargs['pk'])
        devices = system.devices.all()
        return render(request, self.template_name, {'system': system, 'devices': devices})

    def post(self, request, *args, **kwargs):
        return redirect(reverse('poam:edit-device', kwargs={'pk': request.POST.get('device')}))


class EditDeviceView(LoginRequiredMixin, generic.UpdateView):
    template_name = 'poam/edit-device.html'
    model = Device
    form_class = DeviceForm

    def get_success_url(self):
        url = reverse('poam:edit-system', kwargs={'pk': Device.objects.get(id=self.kwargs['pk']).system.id})
        return url


class UploadPOAMView(LoginRequiredMixin, generic.UpdateView):
    template_name = 'poam/upload-poam.html'
    model = System
    form_class = POAMForm

    def form_valid(self, form):
        data = form.cleaned_data['file']
        chosen_system = form.cleaned_data['name']
        try:
            wb = openpyxl.load_workbook(data)
            sheet = wb.active
        except:
            messages.error(self.request, 'Wrong File Type, Zac!')
            return redirect(reverse("poam:upload-poam", kwargs={'pk': self.kwargs['pk']}))
        else:
            create_date = sheet['B1'].value
            update_date = sheet['B2'].value
            dod_component = sheet['B3'].value
            system_name = sheet['B4'].value
            dod_it_resource_number = sheet['B5'].value

            system_type = sheet['H1'].value
            poc_name = sheet['H3'].value
            poc_email = sheet['H4'].value
            poc_phone = sheet['H5'].value

            if system_name != chosen_system:
                messages.error(self.request, 'I think you chose the wrong system! Make sure the system name in your file (cell B4) matches exactly')
                return 0

            try:
                system = System.objects.get(name=system_name)
            except System.DoesNotExist:
                messages.error(self.request, 'Looks like that sysem is not in the database!')
                return 0
            else:
                System.objects.filter(name=system_name).update(update_date=update_date, dod_it_resource_number=dod_it_resource_number)
                system = System.objects.get(name=system_name)

            for row in range(7, sheet.max_row + 1):
                if sheet['A{}'.format(row)].value == None:
                    break

                weakness = sheet['A{}'.format(row)].value

                # Gets everything before Title: - vuln_id
                vuln_id = weakness.split("Title:")[0]
                vuln_id = vuln_id.rstrip()

                # Gets everything after Title:
                weakness = weakness.split("Title:")[1]

                # Gets everything before Description: - title
                title = weakness.split("Description:")[0]
                title = title.rstrip()

                # Gets everything after Description
                weakness = weakness.split("Description:")[1]

                # Gets everything after Devices Affected:
                devices_affected = weakness.split("Devices Affected:\n")[1]
                devices_affected = devices_affected.rstrip()
                devices_affected = devices_affected.rsplit('\n')

                devices = []

                devices_in_db = True

                for device_affected in devices_affected:
                    try:
                        Device.objects.get(name=device_affected)
                    except Device.DoesNotExist:
                        messages.error(self.request, 'Device {} is not in the database! Please go back and add it before continuing'.format(device_affected))
                        devices_in_db = False
                    else:
                        device = Device.objects.get(name=device_affected)
                        devices.append(device)

                if not devices_in_db:
                    break

                # Gets everything before Devices Affected: - Description
                description = weakness.split("Devices Affected:")[0]
                description = description.rstrip()

                raw_severity = sheet['B{}'.format(row)].value
                security_control = sheet['C{}'.format(row)].value
                mitigated_severity = sheet['D{}'.format(row)].value
                mitigation = sheet['E{}'.format(row)].value
                poc = sheet['F{}'.format(row)].value.splitlines()
                poc_name = poc[0]
                poc_email = poc[1]
                # poc_phone = poc[2]
                try:
                    point_of_contact = PointOfContact.objects.get(name=poc_name)
                except PointOfContact.DoesNotExist:
                    PointOfContact.objects.create(name=poc_name, email=poc_email, phone='')
                else:
                    point_of_contact = PointOfContact.objects.get(name=poc_name)
                resources_required = sheet['G{}'.format(row)].value
                scheduled_completion_date = sheet['H{}'.format(row)].value
                milestone_changes = sheet['I{}'.format(row)].value
                siw = sheet['J{}'.format(row)].value.splitlines()
                source_identifying_event = siw[0].split('1.')[1]
                source_identifying_tool = siw[1].split('2.')[1]
                source_identifying_date = siw[2].split('3.')[1]
                status = sheet['K{}'.format(row)].value

                comments = sheet['L{}'.format(row)].value.splitlines()

                try:
                    finding_details = comments[2].split('Finding Details: ')[1]
                except IndexError:
                    finding_details = ''

                try:
                    cvss_base_score = comments[5].split('CVSS2 Base Score - ')[1]
                except IndexError:
                    cvss_base_score = ''

                try:
                    cvss_temporal_score = comments[6].split('CVSS2 Temporal Score - ')[1]
                except IndexError:
                    cvss_temporal_score = ''

                try:
                    cvss_vector = comments[7].split('CVSS2 Vector - ')[1]
                except IndexError:
                    cvss_vector = ''

                try:
                    cvss_temporal_vector = comments[8].split('CVSS2 Temporal Vector - ')[1]
                except IndexError:
                    cvss_temporal_vector = ''

                try:
                    cvss3_base_score = comments[9].split('CVSS3 Base Score - ')[1]
                except IndexError:
                    cvss3_base_score = ''

                try:
                    cvss3_vector = comments[10].split('CVSS3 Vector - ')[1]
                except IndexError:
                    cvss3_vector = ''

                try:
                    comments = comments[0].split('Comments:  ')[1]
                except IndexError:
                    comments = ''

                try:
                    Weakness.objects.get(vuln_id=VulnId.objects.get(vuln_id=vuln_id), system=system)
                except Weakness.DoesNotExist:
                    data_dict = {'title': title, 'description': description, 'status': status, 'comments': comments, 'finding_details': finding_details, 'raw_severity': raw_severity, 'source_identifying_event': source_identifying_event, 'source_identifying_tool': source_identifying_tool, 'source_identifying_date': source_identifying_date, 'vuln_id': VulnId.objects.get(vuln_id=vuln_id).id, 'system': system.id, 'point_of_contact': point_of_contact.id, 'devices': devices, 'security_control': security_control, 'mitigated_severity': mitigated_severity, 'mitigation': mitigation, 'resources_required': resources_required, 'scheduled_completion_date': scheduled_completion_date, 'milestone_changes': milestone_changes, 'cvss_base_score': cvss_base_score, 'cvss_temporal_score': cvss_temporal_score, 'cvss_vector': cvss_vector, 'cvss_temporal_vector': cvss_temporal_vector, 'cvss3_base_score': cvss3_base_score, 'cvss3_vector': cvss3_vector}
                    form = WeaknessModelForm(data_dict)
                    form.save()
                else:
                    Weakness.objects.filter(vuln_id=VulnId.objects.get(vuln_id=vuln_id)).update(security_control=security_control, mitigated_severity=mitigated_severity, mitigation=mitigation, resources_required=resources_required, scheduled_completion_date=scheduled_completion_date, milestone_changes=milestone_changes, status=status, comments=comments)

        return redirect(reverse("poam:upload-poam", kwargs={'pk': self.kwargs['pk']}))


class UploadDeviceView(LoginRequiredMixin, generic.CreateView):
    template_name = 'poam/upload_device.html'
    model = Device
    form_class = DeviceUploadForm

    def get_context_data(self, **kwargs):
        context = super(UploadDeviceView, self).get_context_data(**kwargs)
        context['system'] = System.objects.get(id=self.kwargs['pk'])
        return context

    def form_invalid(self, form):
        print(form.errors)

    def form_valid(self, form):
        print(form)
